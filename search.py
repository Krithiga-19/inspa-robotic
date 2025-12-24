import tkinter as tk
from tkinter import messagebox, filedialog
import openpyxl
import os
import re
import shutil

# Global variable to store the photo path
photo_filename = ""

def upload_photo():
    global photo_filename
    file_path = filedialog.askopenfilename(title="Select a Photo", filetypes=[("Image Files", "*.jpg;*.jpeg;*.png")])
    if file_path:
        if not os.path.exists("photos"):
            os.makedirs("photos")
        filename = os.path.basename(file_path)
        destination = os.path.join("photos", filename)
        shutil.copy(file_path, destination)
        photo_filename = filename
        messagebox.showinfo("Success", "Photo uploaded successfully!")

def validate_data():
    phone_pattern = r"^\d{10}$"
    email_pattern = r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$"
    if not entries[0].get():
        messagebox.showerror("Error", "Full Name is required.")
        return False
    if not entries[4].get().isdigit() or not re.match(phone_pattern, entries[4].get()):
        messagebox.showerror("Error", "Enter a valid 10-digit Phone Number.")
        return False
    if not re.match(email_pattern, entries[5].get()):
        messagebox.showerror("Error", "Enter a valid Email Address.")
        return False
    if not photo_filename:
        messagebox.showerror("Error", "Please upload a photo.")
        return False
    return True

def save_data():
    if not validate_data():
        return
    data = [entries[i].get() for i in range(11)] + [class_var.get(), photo_filename]
    filename = "student_data.xlsx"
    if not os.path.exists(filename):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        headers = ["Full Name", "Date of Birth", "Gender", "Class", "Religion", "Address", "Phone Number", "Email", "Nationality", "Guardian's Name", "Guardian's Contact", "Previous School", "Admission Date", "Photo"]
        sheet.append(headers)
        workbook.save(filename)
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    sheet.append(data)
    workbook.save(filename)
    workbook.close()
    messagebox.showinfo("Success", "Student details saved successfully!")
    reset_fields()

def reset_fields():
    global photo_filename
    for entry in entries:
        entry.delete(0, tk.END)
    gender_var.set("None")
    class_var.set("Select class")
    photo_filename = ""

def search_student():
    search_key = search_entry.get().strip()
    if not search_key:
        messagebox.showerror("Error", "Please enter a name or phone number to search.")
        return
    filename = "student_data.xlsx"
    if not os.path.exists(filename):
        messagebox.showerror("Error", "No student data found!")
        return
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    found = False
    for row in sheet.iter_rows(values_only=True):
        if search_key in row[:1] or search_key == row[6]:
            for i in range(11):
                entries[i].delete(0, tk.END)
                entries[i].insert(0, row[i])
            gender_var.set(row[2])
            class_var.set(row[3])
            global photo_filename
            photo_filename = row[13]
            found = True
            break
    workbook.close()
    if found:
        messagebox.showinfo("Success", "Student record found!")
    else:
        messagebox.showerror("Error", "No matching student found.")

# GUI Setup
root = tk.Tk()
root.title("Student Registration System")
root.geometry("750x750")
root.configure(bg="purple")

# Header
search_frame = tk.Frame(root, bg="purple")
search_frame.pack(pady=5)
tk.Label(search_frame, text="Search Student (Name/Phone):", bg="purple", fg="white").pack(side=tk.LEFT)
search_entry = tk.Entry(search_frame, width=30)
search_entry.pack(side=tk.LEFT, padx=5)
tk.Button(search_frame, text="Search", command=search_student, bg="blue", fg="white").pack(side=tk.LEFT)

form_frame = tk.Frame(root, bg="purple")
form_frame.pack(pady=10)
fields = ["Full Name", "Date of Birth", "Gender", "Class", "Religion", "Address", "Phone Number", "Email", "Nationality", "Guardian's Name", "Guardian's Contact", "Previous School", "Admission Date", "Upload Photo"]
entries = []
gender_var = tk.StringVar(value="None")
class_var = tk.StringVar(value="Select class")
for idx, text in enumerate(fields):
    tk.Label(form_frame, text=text + ":", fg="white", bg="purple").grid(row=idx, column=0, sticky="w", padx=10, pady=5)
    if text == "Gender":
        tk.Radiobutton(form_frame, text="Male", variable=gender_var, value="Male", bg="purple", fg="white").grid(row=idx, column=1, sticky="w")
        tk.Radiobutton(form_frame, text="Female", variable=gender_var, value="Female", bg="purple", fg="white").grid(row=idx, column=2, sticky="w")
    elif text == "Class":
        class_dropdown = tk.OptionMenu(form_frame, class_var, "B.Tech", "B.Com", "Arts and Science", "Agricultural Science")
        class_dropdown.grid(row=idx, column=1, padx=10, pady=5)
    elif text == "Upload Photo":
        tk.Button(form_frame, text="Choose File", command=upload_photo, bg="blue", fg="white").grid(row=idx, column=1, padx=10, pady=5)
    else:
        entry = tk.Entry(form_frame, width=30)
        entry.grid(row=idx, column=1, padx=10, pady=5)
        entries.append(entry)

button_frame = tk.Frame(root, bg="purple")
button_frame.pack(pady=10)
tk.Button(button_frame, text="Save & Next", command=save_data, bg="green", fg="white", width=15).grid(row=0, column=0, padx=5)
tk.Button(button_frame, text="Reset", command=reset_fields, bg="red", fg="white", width=10).grid(row=0, column=1, padx=5)
tk.Button(button_frame, text="Exit", command=root.destroy, bg="gray", fg="white", width=10).grid(row=0, column=2, padx=5)

root.mainloop()
